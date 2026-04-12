"""
Excel DCF Builder — export and import a fully-formula-driven DCF workbook.

Requires: openpyxl  (pip install openpyxl)

Sheets produced
───────────────
  Inputs       — editable assumptions (yellow cells)
  Projections  — year-by-year FCF build driven by Inputs formulas
  DCF          — PV calculations, terminal value, implied share price
  Sensitivity  — 2-D WACC × terminal-growth table (data-table readiness)

After opening the workbook in Excel all cells auto-calculate.
The Python `import_dcf_from_excel` function reads Inputs back for re-use.
"""

from __future__ import annotations

import warnings
from pathlib import Path
from typing import Optional, List, Union

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers,  # noqa: F401 (not used directly, kept for readability)
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ── Palette ──────────────────────────────────────────────────────────────────
_C_HEADER      = "4B2C82"   # Deep Purple
_C_HEADER_FG   = "FFFFFF"   # White
_C_INPUT       = "FFF2CC"   # Light Yellow
_C_FORMULA     = "F3E5F5"   # Very Light Purple/Lavender
_C_OUTPUT      = "E1BEE7"   # Light Purple/Pinkish
_C_SECTION     = "D1C4E9"   # Medium Light Purple
_C_SECTION_FG  = "000000"   # Black
_C_BORDER      = "BCBCBC"

_PCT_FMT  = "0.00%"
_CURR_FMT = '#,##0.00'
_INT_FMT  = '#,##0'
_DEC_FMT  = '#,##0.0000'


def _require_openpyxl() -> None:
    if not _OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with:  pip install openpyxl"
        )


def _border(color: str = _C_BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row: int, col: int, value) -> None:
    # Shift B2: row+1, col+1
    cell = ws.cell(row=row + 1, column=col + 1, value=value)
    cell.font = Font(bold=True, color=_C_HEADER_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=_C_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def _input_cell(ws, row: int, col: int, value, num_format: str = _CURR_FMT) -> openpyxl.cell.Cell:
    cell = ws.cell(row=row + 1, column=col + 1, value=value)
    cell.fill = PatternFill("solid", fgColor=_C_INPUT)
    cell.number_format = num_format
    cell.border = _border()
    cell.alignment = Alignment(horizontal="right")
    return cell


def _formula_cell(ws, row: int, col: int, formula: str, num_format: str = _CURR_FMT) -> openpyxl.cell.Cell:
    cell = ws.cell(row=row + 1, column=col + 1, value=formula)
    cell.fill = PatternFill("solid", fgColor=_C_FORMULA)
    cell.number_format = num_format
    cell.border = _border()
    cell.alignment = Alignment(horizontal="right")
    return cell


def _output_cell(ws, row: int, col: int, formula: str, num_format: str = _CURR_FMT) -> openpyxl.cell.Cell:
    cell = ws.cell(row=row + 1, column=col + 1, value=formula)
    cell.fill = PatternFill("solid", fgColor=_C_OUTPUT)
    cell.number_format = num_format
    cell.font = Font(bold=True)
    cell.border = _border()
    cell.alignment = Alignment(horizontal="right")
    return cell


def _label(ws, row: int, col: int, text: str, bold: bool = False, section: bool = False) -> None:
    cell = ws.cell(row=row + 1, column=col + 1, value=text)
    cell.font = Font(bold=bold or section, size=10, color=_C_SECTION_FG if section else "000000")
    cell.border = _border()
    if section:
        cell.fill = PatternFill("solid", fgColor=_C_SECTION)
    cell.alignment = Alignment(vertical="center")

def _section_row(ws, row: int, ncols: int, text: str) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row + 1, column=c + 1)
        cell.fill = PatternFill("solid", fgColor=_C_SECTION)
        cell.border = _border()
    ws.cell(row=row + 1, column=2, value=text).font = Font(bold=True, size=10, color=_C_SECTION_FG)


def _get_ref(row: int, col: int, sheet: Optional[str] = None) -> str:
    """Logical (row, col) to Excel reference (B2 start)."""
    c_let = get_column_letter(col + 1)
    r_num = row + 1
    if sheet:
        # Quote sheet name if it has spaces
        s_name = f"'{sheet}'" if " " in sheet else sheet
        return f"{s_name}!{c_let}{r_num}"
    return f"{c_let}{r_num}"


# ─────────────────────────────────────────────────────────────────────────────
# As Is sheet
# ─────────────────────────────────────────────────────────────────────────────

def _build_dashboard_sheet(
    wb: Workbook,
    *,
    company_name: str,
    current_price: float,
    base_revenue: float,
    shares_outstanding: float,
    net_debt: float,
) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25

    # Row 1 logical -> B2 in Excel
    row = 1
    _header_cell(ws, row, 1, f"Company: {company_name}")
    _header_cell(ws, row, 2, "Company Dashboard")
    row += 2

    _section_row(ws, row, 2, "1. Market & Operating Assumptions")
    row += 1
    _label(ws, row, 1, "Current Share Price")
    _input_cell(ws, row, 2, current_price, _CURR_FMT)
    row += 1
    _label(ws, row, 1, "Base Revenue (Year 0)")
    _input_cell(ws, row, 2, base_revenue, _CURR_FMT)
    row += 2

    _section_row(ws, row, 2, "2. Capital Structure (Historical/Current)")
    row += 1
    _label(ws, row, 1, "Shares Outstanding")
    _input_cell(ws, row, 2, shares_outstanding, _INT_FMT)
    row += 1
    _label(ws, row, 1, "Total Debt")
    _input_cell(ws, row, 2, net_debt * 1.2, _CURR_FMT) # Proxy for Gross Debt
    row += 1
    _label(ws, row, 1, "Cash & Equivalents")
    _input_cell(ws, row, 2, net_debt * 0.2, _CURR_FMT) # Proxy for Cash
    row += 1
    _label(ws, row, 1, "Net Debt (Debt - Cash)", bold=True)
    _output_cell(ws, row, 2, f"={_get_ref(row-2, 2)} - {_get_ref(row-1, 2)}", _CURR_FMT)
    row += 2


# ─────────────────────────────────────────────────────────────────────────────
# WACC sheet
# ─────────────────────────────────────────────────────────────────────────────

def _build_wacc_sheet(
    wb: Workbook,
    *,
    risk_free_rate: float,
    beta: float,
    market_risk_premium: float,
    cost_of_debt_pre_tax: float,
    tax_rate: float,
) -> None:
    ws = wb.create_sheet("WACC")
    ws.sheet_view.showGridLines = False
    
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15

    row = 1
    _header_cell(ws, row, 1, "Cost of Capital Calculation")
    _header_cell(ws, row, 2, "Formula/Reference")
    _header_cell(ws, row, 3, "Units")
    row += 2

    _section_row(ws, row, 3, "Part 1: Cost of Equity (CAPM)")
    row += 1
    
    _label(ws, row, 1, "Risk-Free Rate")
    _input_cell(ws, row, 2, risk_free_rate, _PCT_FMT)
    _label(ws, row, 3, "%")
    row += 1

    _label(ws, row, 1, "Beta")
    _input_cell(ws, row, 2, beta, "0.00")
    _label(ws, row, 3, "x")
    row += 1

    _label(ws, row, 1, "Market Risk Premium")
    _input_cell(ws, row, 2, market_risk_premium, _PCT_FMT)
    _label(ws, row, 3, "%")
    row += 1

    _label(ws, row, 1, "Cost of Equity (Ke)", bold=True)
    ke_formula = f"={_get_ref(row-3, 2)} + ({_get_ref(row-2, 2)} * {_get_ref(row-1, 2)})"
    _output_cell(ws, row, 2, ke_formula, _PCT_FMT)
    _label(ws, row, 3, "%")
    row += 2

    _section_row(ws, row, 3, "Part 2: Cost of Debt")
    row += 1

    _label(ws, row, 1, "Pre-tax Cost of Debt")
    _input_cell(ws, row, 2, cost_of_debt_pre_tax, _PCT_FMT)
    _label(ws, row, 3, "%")
    row += 1

    _label(ws, row, 1, "Marginal Tax Rate")
    _input_cell(ws, row, 2, tax_rate, _PCT_FMT)
    _label(ws, row, 3, "%")
    row += 1

    _label(ws, row, 1, "After-tax Cost of Debt (Kd)", bold=True)
    kd_formula = f"={_get_ref(row-2, 2)} * (1 - {_get_ref(row-1, 2)})"
    _output_cell(ws, row, 2, kd_formula, _PCT_FMT)
    _label(ws, row, 3, "%")
    row += 2

    _section_row(ws, row, 3, "Part 3: Capital Structure Weights (Market Based)")
    row += 1
    
    _label(ws, row, 1, "Current Share Price")
    # Dashboard logical row 4 (Excel C5)
    _formula_cell(ws, row, 2, f"={_get_ref(4, 2, 'Dashboard')}", _CURR_FMT)
    price_ref = _get_ref(row, 2)
    row += 1

    _label(ws, row, 1, "Shares Outstanding")
    # Dashboard logical row 7 (Excel C8)
    _formula_cell(ws, row, 2, f"={_get_ref(7, 2, 'Dashboard')}", _INT_FMT)
    shares_ref = _get_ref(row, 2)
    row += 1

    _label(ws, row, 1, "Market Value of Equity (E)")
    _formula_cell(ws, row, 2, f"={price_ref} * {shares_ref}", _CURR_FMT)
    mve_ref = _get_ref(row, 2)
    row += 1

    _label(ws, row, 1, "Market Value of Net Debt (D)")
    # Dashboard logical row 10 (Excel C11)
    _formula_cell(ws, row, 2, f"={_get_ref(10, 2, 'Dashboard')}", _CURR_FMT)
    debt_ref = _get_ref(row, 2)
    row += 1

    _label(ws, row, 1, "Total Capital (V = E + D)")
    _formula_cell(ws, row, 2, f"={mve_ref} + {debt_ref}", _CURR_FMT)
    total_val_ref = _get_ref(row, 2)
    row += 2

    _label(ws, row, 1, "Weight of Equity (We)", bold=True)
    _output_cell(ws, row, 2, f"={mve_ref} / {total_val_ref}", _PCT_FMT)
    we_ref = _get_ref(row, 2)
    row += 1

    _label(ws, row, 1, "Weight of Debt (Wd)", bold=True)
    _output_cell(ws, row, 2, f"={debt_ref} / {total_val_ref}", _PCT_FMT)
    wd_ref = _get_ref(row, 2)
    row += 2

    _section_row(ws, row, 3, "Part 4: Estimating WACC")
    row += 1
    _label(ws, row, 1, "WACC", bold=True)
    # Ke * We + Kd * Wd
    ke_ref = _get_ref(7, 2)  # Logical row 7 is Ke
    kd_ref = _get_ref(12, 2) # Logical row 12 is Kd
    wacc_formula = f"=({ke_ref} * {we_ref}) + ({kd_ref} * {wd_ref})"
    _output_cell(ws, row, 2, wacc_formula, _PCT_FMT)
    row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Projections sheet
# ─────────────────────────────────────────────────────────────────────────────
# Projection sheet
# ─────────────────────────────────────────────────────────────────────────────

def _build_valuation_sheet(
    wb: Workbook,
    *,
    years: int,
    revenue_growth_rates: List[float],
    ebit_margins: List[float],
    capex_pct: List[float],
    da_pct: List[float],
    nwc_change_pct: List[float],
    terminal_growth: float,
) -> None:
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 40

    # PART 1: FCF FORECAST
    row = 1
    _header_cell(ws, row, 1, "Integrated DCF Valuation Model")
    for yr in range(1, years + 1):
        _header_cell(ws, row, 1 + yr, f"Year {yr}")
        ws.column_dimensions[get_column_letter(2 + yr)].width = 18
    row += 2

    _section_row(ws, row, 1 + years, "Section A: Growth & Margin Assumptions")
    row += 1
    
    assumptions = [
        ("Revenue Growth Rate", revenue_growth_rates, _PCT_FMT),
        ("EBIT Margin", ebit_margins, _PCT_FMT),
        ("CapEx (% Revenue)", capex_pct, _PCT_FMT),
        ("D&A (% Revenue)", da_pct, _PCT_FMT),
        ("ΔWorking Capital (% Revenue)", nwc_change_pct, _PCT_FMT),
    ]
    
    assumption_rows = {}
    for label_text, values, fmt in assumptions:
        _label(ws, row, 1, label_text)
        assumption_rows[label_text] = row
        for j, val in enumerate(values):
            _input_cell(ws, row, 2 + j, val, fmt)
        row += 1
    row += 2

    _section_row(ws, row, 1 + years, "Section B: Free Cash Flow (FCF) Build")
    row += 1
    
    # 1. Revenue
    _label(ws, row, 1, "Revenue")
    rev_row = row
    for yr in range(1, years + 1):
        growth_ref = _get_ref(assumption_rows['Revenue Growth Rate'], 1 + yr)
        if yr == 1:
            # Dashboard C5 (4, 2)
            formula = f"={_get_ref(4, 2, 'Dashboard')} * (1 + {growth_ref})"
        else:
            prev_rev = _get_ref(rev_row, yr)
            formula = f"={prev_rev} * (1 + {growth_ref})"
        _formula_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 1

    # 2. EBITDA
    _label(ws, row, 1, "EBITDA")
    ebitda_row = row
    for yr in range(1, years + 1):
        rev_ref = _get_ref(rev_row, 1 + yr)
        ebit_margin_ref = _get_ref(assumption_rows['EBIT Margin'], 1 + yr)
        da_margin_ref   = _get_ref(assumption_rows['D&A (% Revenue)'], 1 + yr)
        formula = f"={rev_ref} * ({ebit_margin_ref} + {da_margin_ref})"
        _formula_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 1

    # 3. D&A (Negative)
    _label(ws, row, 1, "(-) Depreciation & Amortization")
    da_row = row
    for yr in range(1, years + 1):
        rev_ref = _get_ref(rev_row, 1 + yr)
        da_margin_ref = _get_ref(assumption_rows['D&A (% Revenue)'], 1 + yr)
        formula = f"=-({rev_ref} * {da_margin_ref})"
        _formula_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 1

    # 4. EBIT
    _label(ws, row, 1, "EBIT (Operating Income)", bold=True)
    ebit_row = row
    for yr in range(1, years + 1):
        formula = f"={_get_ref(ebitda_row, 1 + yr)} + {_get_ref(da_row, 1 + yr)}"
        _output_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 1

    # 5. Taxes
    _label(ws, row, 1, "(-) Cash Taxes")
    tax_row = row
    # WACC sheet logical row 10 (Excel C11) is Marginal Tax Rate
    tax_rate_ref = _get_ref(10, 2, "WACC")
    for yr in range(1, years + 1):
        formula = f"=-({_get_ref(ebit_row, 1 + yr)} * {tax_rate_ref})"
        _formula_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 1

    # 6. NOPAT
    _label(ws, row, 1, "NOPAT", bold=True)
    nopat_row = row
    for yr in range(1, years + 1):
        formula = f"={_get_ref(ebit_row, 1 + yr)} + {_get_ref(tax_row, 1 + yr)}"
        _output_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 1

    _label(ws, row, 1, "(+) Depreciation & Amortization")
    for yr in range(1, years + 1):
        formula = f"=-{_get_ref(da_row, 1 + yr)}"
        _formula_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 1

    _label(ws, row, 1, "(-) Capital Expenditures")
    capex_row = row
    for yr in range(1, years + 1):
        rev_ref = _get_ref(rev_row, 1 + yr)
        # We need assumption row for capex pct
        capex_pct_ref = _get_ref(assumption_rows['CapEx (% Revenue)'], 1 + yr)
        formula = f"=-({rev_ref} * {capex_pct_ref})"
        _formula_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 1

    _label(ws, row, 1, "(-) ΔWorking Capital")
    nwc_row = row
    for yr in range(1, years + 1):
        rev_ref = _get_ref(rev_row, 1 + yr)
        nwc_pct_ref = _get_ref(assumption_rows['ΔWorking Capital (% Revenue)'], 1 + yr)
        formula = f"=-({rev_ref} * {nwc_pct_ref})"
        _formula_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 1

    _label(ws, row, 1, "Unlevered Free Cash Flow (UFCF)", bold=True)
    fcf_row = row
    for yr in range(1, years + 1):
        # Sum from NOPAT (r-3) to NWC (r-1)
        # NOPAT is row 17 (logical), Add DA is 18, CapEx 19, NWC 20. FCF 21.
        # Wait, let's just sum the relevant rows.
        formula = f"=SUM({_get_ref(nopat_row, 1+yr)}:{_get_ref(nwc_row, 1+yr)})"
        _output_cell(ws, row, 1 + yr, formula, _CURR_FMT)
    row += 3

    # PART 2: DISCOUNTED CASH FLOW ANALYSIS
    _section_row(ws, row, 2, "Section C: Discounted Cash Flow (DCF) Analysis")
    row += 1
    
    _label(ws, row, 1, "Terminal Growth Rate (g)")
    _input_cell(ws, row, 2, terminal_growth, _PCT_FMT)
    tg_ref = _get_ref(row, 2)
    row += 1
    _label(ws, row, 1, "WACC (from WACC sheet)")
    # Logical row 25 in WACC is official WACC (Excel C26)
    _formula_cell(ws, row, 2, f"={_get_ref(25, 2, 'WACC')}", _PCT_FMT)
    wacc_ref = _get_ref(row, 2)
    row += 2

    # PV Calculation Table
    _header_cell(ws, row, 1, "Step")
    _header_cell(ws, row, 2, "Formula Part")
    _header_cell(ws, row, 3, "Result")
    row += 1

    _label(ws, row, 1, "Discounted Cash Flow Table")
    # Headers for PV table
    ws.cell(row=row + 1, column=3 + 1, value="Year").font = Font(bold=True)
    ws.cell(row=row + 1, column=4 + 1, value="FCF").font = Font(bold=True)
    ws.cell(row=row + 1, column=5 + 1, value="DF").font = Font(bold=True)
    ws.cell(row=row + 1, column=6 + 1, value="PV").font = Font(bold=True)
    row += 1

    pv_start_row = row
    for yr in range(1, years + 1):
        _label(ws, row, 3, f"Year {yr}")
        # FCF from above
        fcf_ref = _get_ref(fcf_row, 1 + yr)
        _formula_cell(ws, row, 4, f"={fcf_ref}", _CURR_FMT)
        # DF = 1 / (1+WACC)^yr
        _formula_cell(ws, row, 5, f"=1 / (1 + {wacc_ref})^{yr}", "0.0000")
        # PV = FCF * DF
        _formula_cell(ws, row, 6, f"={_get_ref(row, 4)} * {_get_ref(row, 5)}", _CURR_FMT)
        row += 1
    pv_end_row = row - 1
    row += 2

    _section_row(ws, row, 2, "Section D: Intrinsic Value Attribution")
    row += 1
    _label(ws, row, 1, "Enterprise Value (EV) Build", bold=True)
    row += 1
    
    _label(ws, row, 1, "PV of Explicit Forecast")
    sum_range = f"{get_column_letter(6+1)}{pv_start_row+1}:{get_column_letter(6+1)}{pv_end_row+1}"
    _output_cell(ws, row, 2, f"=SUM({sum_range})", _CURR_FMT)
    pv_expl_ref = _get_ref(row, 2)
    row += 1

    _label(ws, row, 1, "Terminal Value (TV)")
    last_fcf_ref = _get_ref(pv_end_row, 4)
    # TV_n = FCF_n * (1+g) / (WACC - g)
    tv_formula = f"={last_fcf_ref} * (1 + {tg_ref}) / ({wacc_ref} - {tg_ref})"
    _formula_cell(ws, row, 2, tv_formula, _CURR_FMT)
    tv_val_ref = _get_ref(row, 2)
    row += 1

    _label(ws, row, 1, "PV of Terminal Value")
    last_df_ref = _get_ref(pv_end_row, 5)
    _formula_cell(ws, row, 2, f"={tv_val_ref} * {last_df_ref}", _CURR_FMT)
    pv_tv_ref = _get_ref(row, 2)
    row += 2

    _label(ws, row, 1, "TOTAL ENTERPRISE VALUE", bold=True)
    _output_cell(ws, row, 2, f"={pv_expl_ref} + {pv_tv_ref}", _CURR_FMT)
    ev_total_ref = _get_ref(row, 2)
    row += 1

    _label(ws, row, 1, "(-) Net Debt")
    # Dashboard logical row 10 (Excel C11)
    _formula_cell(ws, row, 2, f"={_get_ref(10, 2, 'Dashboard')}", _CURR_FMT)
    debt_final_ref = _get_ref(row, 2)
    row += 2

    _label(ws, row, 1, "TOTAL EQUITY VALUE", bold=True)
    _output_cell(ws, row, 2, f"={ev_total_ref} - {debt_final_ref}", _CURR_FMT)
    equity_val_ref = _get_ref(row, 2)
    row += 1

    _label(ws, row, 1, "(/) Shares Outstanding")
    # Dashboard logical row 7 (Excel C8)
    _formula_cell(ws, row, 2, f"={_get_ref(7, 2, 'Dashboard')}", _INT_FMT)
    shares_final_ref = _get_ref(row, 2)
    row += 2

    _label(ws, row, 1, "IMPLIED SHARE PRICE", bold=True)
    _output_cell(ws, row, 2, f"={equity_val_ref} / {shares_final_ref}", _CURR_FMT)
    final_price_cell = ws.cell(row=row + 1, column=2 + 1)
    final_price_cell.font = Font(bold=True, size=14, color="4B2C82") # Purple highlight
    row += 2

    ws.freeze_panes = "C3"
    # Highlight the final result
    final_cell = ws.cell(row=row + 1, column=3)
    final_cell.font = Font(bold=True, size=12, color=_C_HEADER)
    row += 2


def _build_sensitivity_sheet(
    wb: Workbook,
    *,
    base_fcf_vals: List[float],
    sensitivity_waccs: List[float],
    sensitivity_tgs: List[float],
    shares_outstanding: float,
    net_debt: float,
) -> None:
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 15

    # Headers
    row = 1
    _header_cell(ws, row, 1, "Sensitivity: Implied Share Price")
    # Use C2-I2 as well to make it look framed
    for j in range(2, 9):
        _header_cell(ws, row, j, "")
    row += 2

    # Starting B4
    _label(ws, row, 1, "WACC \\ Terminal g", bold=True, section=True)
    
    for j, g in enumerate(sensitivity_tgs, start=2):
        _header_cell(ws, row, j, g)
        ws.cell(row=row+1, column=j+1).number_format = _PCT_FMT
        ws.column_dimensions[get_column_letter(j + 1)].width = 14
    row += 1
    from .equities import dcf_valuation
  # local import to avoid circular

    for i, w in enumerate(sensitivity_waccs, start=row + 1):
        # logical i, col 1 -> Excel row i+1, col 2
        row_c = ws.cell(row=i + 1, column=2, value=w)
        row_c.number_format = _PCT_FMT
        row_c.font = Font(bold=True)
        row_c.fill = PatternFill("solid", fgColor=_C_SECTION)
        row_c.font = Font(bold=True, color=_C_SECTION_FG)
        row_c.border = _border()

        for j, g in enumerate(sensitivity_tgs, start=2):
            if w <= g:
                val = None
            else:
                result = dcf_valuation(
                    base_fcf_vals, w, g,
                    shares_outstanding=shares_outstanding,
                    net_debt=net_debt,
                )
                val = result["price_per_share"]
            # logical i, col j -> Excel row i+1, col j+1
            c = ws.cell(row=i + 1, column=j + 1, value=val)
            c.number_format = _CURR_FMT
            c.border = _border()

    # Colour scale
    # start_excel_row = row + 1 + 1 = row + 2
    # start_excel_col = 3 (C)
    # end_excel_row = row + len(sensitivity_waccs) + 1
    # end_excel_col = 1 + len(sensitivity_tgs) + 1
    s_row = row + 2
    e_row = row + len(sensitivity_waccs) + 1
    s_col = 3
    e_col = len(sensitivity_tgs) + 2
    
    tbl_ref = f"{get_column_letter(s_col)}{s_row}:{get_column_letter(e_col)}{e_row}"
    ws.conditional_formatting.add(
        tbl_ref,
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )
    ws.freeze_panes = "C5"


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def export_dcf_to_excel(
    path: Union[str, Path],
    *,
    company_name: str = "Company",
    years: int = 5,
    current_price: float = 100.0,
    base_revenue: float = 1_000.0,
    revenue_growth_rates: Optional[List[float]] = None,
    ebit_margins: Optional[List[float]] = None,
    tax_rate: float = 0.25,
    capex_pct: Optional[List[float]] = None,
    da_pct: Optional[List[float]] = None,
    nwc_change_pct: Optional[List[float]] = None,
    # CAPM / WACC inputs
    risk_free_rate: Union[float, List[float]] = 0.04,
    beta: float = 1.0,
    market_risk_premium: float = 0.06,
    cost_of_debt_pre_tax: float = 0.06,
    terminal_growth: float = 0.025,
    shares_outstanding: float = 100.0,
    net_debt: float = 0.0,
    sensitivity_waccs: Optional[List[float]] = None,
    sensitivity_tgs: Optional[List[float]] = None,
) -> Path:
    """
    Build and save a fully-formatted, formula-driven DCF workbook.

    Excel sheets:
      - As Is: Current metrics (Revenue, Shares, Debt).
      - WACC: CAPM and WACC calculation with editable inputs.
      - Projection: Forecast build (from Revenue to FCF).
      - DCF: Enterprise and Equity value build using cross-sheet formulas.
      - Sensitivity: 2D table of Share Price vs WACC and Terminal Growth.

    Parameters
    ----------
    path : str or Path
        Output ``.xlsx`` file path.
    company_name : str, default 'Company'
        Name shown in the title.
    years : int, default 5
        Number of forecast years.
    base_revenue : float, default 1000
        Year-0 revenue.
    revenue_growth_rates : list of float, optional
        Per-year revenue growth.
    ebit_margins : list of float, optional
        Per-year EBIT margin.
    tax_rate : float, default 0.25
        Tax rate.
    capex_pct : list of float, optional
        CapEx as % of revenue.
    da_pct : list of float, optional
        D&A as % of revenue.
    nwc_change_pct : list of float, optional
        NWC increase as % of revenue.
    risk_free_rate : float, default 0.04
        Risk-free rate for CAPM.
    beta : float, default 1.0
        Equity beta for CAPM.
    market_risk_premium : float, default 0.06
        Expected market return minus risk-free rate.
    cost_of_debt_pre_tax : float, default 0.06
        Interest rate on debt.
    terminal_growth : float, default 0.025
        Perpetual growth rate.
    shares_outstanding : float, default 100
        Share count.
    net_debt : float, default 0
        Net debt.
    """
    _require_openpyxl()

    # Handle risk_free_rate as list
    if isinstance(risk_free_rate, list):
        # Use the last value for forward-looking WACC
        rf_forward = risk_free_rate[-1]
    else:
        rf_forward = risk_free_rate

    # Defaults
    revenue_growth_rates = revenue_growth_rates or [0.10] * years
    ebit_margins         = ebit_margins         or [0.20] * years
    capex_pct            = capex_pct            or [0.05] * years
    da_pct               = da_pct               or [0.03] * years
    nwc_change_pct       = nwc_change_pct       or [0.02] * years

    # For sensitivity, we still need some "base" WACC
    ke = rf_forward + (beta * market_risk_premium)
    kd_at = cost_of_debt_pre_tax * (1 - tax_rate)
    est_wacc = (0.7 * ke) + (0.3 * kd_at) # proxy for sensitivity range

    sensitivity_waccs = sensitivity_waccs or [round(est_wacc + d * 0.01, 4) for d in range(-3, 4)]
    sensitivity_tgs   = sensitivity_tgs   or [round(terminal_growth + d * 0.005, 4) for d in range(-3, 4)]

    wb = Workbook()
    wb.remove(wb.active)

    # 1. Dashboard
    _build_dashboard_sheet(
        wb,
        company_name=company_name,
        current_price=current_price,
        base_revenue=base_revenue,
        shares_outstanding=shares_outstanding,
        net_debt=net_debt,
    )

    # 2. WACC
    _build_wacc_sheet(
        wb,
        risk_free_rate=rf_forward,
        beta=beta,
        market_risk_premium=market_risk_premium,
        cost_of_debt_pre_tax=cost_of_debt_pre_tax,
        tax_rate=tax_rate,
    )

    # 3. Valuation (Integrated Projection + DCF)
    _build_valuation_sheet(
        wb,
        years=years,
        revenue_growth_rates=revenue_growth_rates,
        ebit_margins=ebit_margins,
        capex_pct=capex_pct,
        da_pct=da_pct,
        nwc_change_pct=nwc_change_pct,
        terminal_growth=terminal_growth,
    )

    # 5. Sensitivity
    # Needs a list of static FCFs to iterate
    # We calculate based on inputs
    rev = base_revenue
    fcf_vals = []
    for i in range(years):
        rev *= (1 + revenue_growth_rates[i])
        ebit = rev * ebit_margins[i]
        nopat = ebit * (1 - tax_rate)
        fcf = nopat + (rev * da_pct[i]) - (rev * capex_pct[i]) - (rev * nwc_change_pct[i])
        fcf_vals.append(fcf)

    _build_sensitivity_sheet(
        wb,
        base_fcf_vals=fcf_vals,
        sensitivity_waccs=sensitivity_waccs,
        sensitivity_tgs=sensitivity_tgs,
        shares_outstanding=shares_outstanding,
        net_debt=net_debt,
    )

    out = Path(path).resolve()
    wb.save(out)
    return out


def import_dcf_from_excel(path: Union[str, Path]) -> dict:
    """
    Read the Input sheets of a DCF workbook and return a parameter dict.
    """
    _require_openpyxl()
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    
    ws_dash = wb["Dashboard"]
    ws_wacc = wb["WACC"]
    ws_val  = wb["Valuation"]

    # Dashboard Static Inputs
    # Price: Logical 4 -> C5
    # Rev: Logical 5 -> C6
    # Shares: Logical 8 -> C9
    # Total Debt: Logical 9 -> C10
    # Cash: Logical 10 -> C11
    price_val    = ws_dash.cell(5, 3).value
    base_rev_val = ws_dash.cell(6, 3).value
    shares_val   = ws_dash.cell(9, 3).value
    debt_gross   = ws_dash.cell(10, 3).value
    cash_val     = ws_dash.cell(11, 3).value

    # WACC Static Inputs
    rf_val   = ws_wacc.cell(5, 3).value
    beta_val = ws_wacc.cell(6, 3).value
    mrp_val  = ws_wacc.cell(7, 3).value
    kd_pre   = ws_wacc.cell(11, 3).value
    tax_val  = ws_wacc.cell(12, 3).value

    # Valuation (Years)
    years_val = 0
    for c in range(3, 50):
        if ws_val.cell(2, c).value: years_val += 1
        else: break

    def _read_val_row(r_logical, n):
        r_excel = r_logical + 1
        return [ws_val.cell(r_excel, 3 + j).value for j in range(n)]

    return {
        "company_name":        str(ws_dash.cell(2, 2).value).replace("Company: ", ""),
        "current_price":       float(price_val or 0),
        "base_revenue":        float(base_rev_val or 0),
        "shares_outstanding":  float(shares_val or 0),
        "gross_debt":          float(debt_gross or 0),
        "cash":                float(cash_val or 0),
        "net_debt":            float((debt_gross or 0) - (cash_val or 0)),
        "risk_free_rate":      float(rf_val or 0),
        "beta":                float(beta_val or 0),
        "market_risk_premium": float(mrp_val or 0),
        "cost_of_debt_pre_tax":float(kd_pre or 0),
        "tax_rate":            float(tax_val or 0),
        "years":               years_val,
        "revenue_growth_rates":[float(v or 0) for v in _read_val_row(4, years_val)],
        "ebit_margins":        [float(v or 0) for v in _read_val_row(5, years_val)],
        "capex_pct":           [float(v or 0) for v in _read_val_row(6, years_val)],
        "da_pct":              [float(v or 0) for v in _read_val_row(7, years_val)],
        "nwc_change_pct":      [float(v or 0) for v in _read_val_row(8, years_val)],
    }
